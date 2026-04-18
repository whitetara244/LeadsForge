# app.py - LeadForge Pro v7.0 (Fixed Role & Priority Detection)
from flask import Flask, request, render_template_string, jsonify, Response
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import re
from datetime import datetime
import csv
import io
import json
from urllib.parse import urlparse
import hashlib
import sqlite3
import logging
from email_validator import validate_email, EmailNotValidError
import phonenumbers
from phonenumbers import carrier, geocoder, timezone
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import secrets
from typing import List, Dict, Any, Optional, Tuple
from collections import defaultdict, Counter
import uuid
import platform
import os

app = Flask(__name__)
CORS(app)

# Configuration
class Config:
    SECRET_KEY = secrets.token_urlsafe(32)
    DATABASE = 'leads.db'
    RATE_LIMIT = "200 per hour"

app.config.from_object(Config)

# Rate limiting
limiter = Limiter(get_remote_address, app=app, default_limits=[Config.RATE_LIMIT], storage_uri="memory://")

# Setup logging
logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

# ==================== Database Setup ====================
def init_db():
    conn = sqlite3.connect(Config.DATABASE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS sessions
                 (id TEXT PRIMARY KEY, status TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS leads
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT, type TEXT, value TEXT,
                  domain TEXT, role TEXT, priority TEXT, confidence_score REAL,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    c.execute('''CREATE TABLE IF NOT EXISTS statistics
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, total_sessions INTEGER DEFAULT 0,
                  total_emails INTEGER DEFAULT 0, total_phones INTEGER DEFAULT 0, date DATE UNIQUE)''')
    conn.commit()
    conn.close()
    app.logger.info("Database initialized")

init_db()

# ==================== Lead Classifier ====================
class LeadClassifier:
    ROLE_PATTERNS = {
        'executive': ['ceo', 'cfo', 'cto', 'coo', 'president', 'chairman', 'director', 'vp', 'vice president', 'owner', 'founder', 'partner', 'executive', 'managing director', 'board'],
        'management': ['manager', 'head', 'lead', 'supervisor', 'coordinator', 'director of', 'senior manager', 'general manager'],
        'sales': ['sales', 'business development', 'bd', 'account executive', 'account manager', 'development', 'partnerships', 'sales rep', 'territory', 'channel'],
        'marketing': ['marketing', 'social media', 'digital', 'content', 'brand', 'pr', 'communications', 'seo', 'sem', 'growth', 'campaign'],
        'technical': ['engineer', 'developer', 'architect', 'technical', 'devops', 'sysadmin', 'it', 'support', 'helpdesk', 'programmer', 'software', 'qa', 'data'],
        'hr': ['hr', 'human resources', 'recruiting', 'talent', 'people', 'hiring', 'recruiter', 'personnel'],
        'finance': ['finance', 'accounting', 'accounts', 'treasury', 'tax', 'audit', 'controller', 'financial', 'budget', 'payroll'],
        'legal': ['legal', 'counsel', 'attorney', 'lawyer', 'compliance', 'regulatory', 'general counsel'],
        'support': ['support', 'help', 'service', 'customer', 'client services', 'customer success', 'technical support'],
        'admin': ['admin', 'office', 'assistant', 'secretary', 'reception', 'front desk', 'clerical', 'coordinator'],
        'generic': ['info', 'contact', 'hello', 'team', 'careers', 'jobs', 'press', 'media', 'webmaster', 'inquiries', 'mail']
    }
    
    ROLE_SCORES = {
        'executive': 100,
        'management': 85,
        'sales': 80,
        'marketing': 75,
        'finance': 70,
        'legal': 70,
        'technical': 65,
        'hr': 55,
        'support': 45,
        'admin': 35,
        'generic': 25,
        'personal': 15
    }
    
    @staticmethod
    def extract_domain(email: str) -> str:
        try:
            if '@' in email:
                domain = email.split('@')[1].lower()
                # Remove common prefixes
                for prefix in ['www.', 'mail.', 'email.']:
                    if domain.startswith(prefix):
                        domain = domain[len(prefix):]
                return domain
        except:
            pass
        return 'unknown'
    
    @staticmethod
    def detect_role(email: str) -> str:
        """Enhanced role detection from email"""
        if '@' not in email:
            return 'generic'
        
        local = email.split('@')[0].lower()
        
        # Check for role patterns first
        for role, patterns in LeadClassifier.ROLE_PATTERNS.items():
            for pattern in patterns:
                if pattern in local:
                    app.logger.info(f"Role detected: {role} from {pattern} in {local}")
                    return role
        
        # Check if it's a personal email (name based)
        parts = re.split(r'[._\-]', local)
        if len(parts) >= 2:
            # Check if it looks like firstname.lastname
            if parts[0].isalpha() and len(parts[0]) >= 2 and parts[1].isalpha() and len(parts[1]) >= 2:
                return 'personal'
            # Check for first initial + last name
            if len(parts[0]) == 1 and parts[0].isalpha() and parts[1].isalpha() and len(parts[1]) >= 2:
                return 'personal'
        
        return 'generic'
    
    @staticmethod
    def get_priority(role: str) -> str:
        """Get priority based on role"""
        if role in ['executive', 'management']:
            return 'high'
        elif role in ['sales', 'marketing', 'finance', 'legal']:
            return 'medium'
        else:
            return 'low'
    
    @staticmethod
    def get_score(role: str) -> int:
        """Get numeric score for role"""
        return LeadClassifier.ROLE_SCORES.get(role, 25)

# ==================== Validator ====================
class Validator:
    @staticmethod
    def validate_email(email: str) -> Tuple[bool, Optional[str], Optional[str]]:
        try:
            validation = validate_email(email, check_deliverability=False)
            normalized = validation.normalized
            # Detect role immediately
            role = LeadClassifier.detect_role(normalized)
            return True, normalized, role
        except:
            return False, None, None
    
    @staticmethod
    def validate_phone(phone: str) -> Tuple[bool, Optional[str]]:
        try:
            cleaned = re.sub(r'[\s\-\(\)\.]', '', phone)
            parsed = phonenumbers.parse(cleaned, None)
            if phonenumbers.is_valid_number(parsed):
                return True, phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
            return False, None
        except:
            return False, None

# ==================== Lead Processor ====================
class LeadProcessor:
    @staticmethod
    def parse_input(text: str) -> List[str]:
        items = re.split(r'[,\n\t ]+', text)
        return [item.strip() for item in items if item.strip() and len(item.strip()) > 3]
    
    @staticmethod
    def parse_csv(content: bytes) -> List[str]:
        try:
            text = content.decode('utf-8', errors='ignore')
            items = []
            csv_reader = csv.reader(io.StringIO(text))
            for row in csv_reader:
                for cell in row:
                    cell = cell.strip()
                    if cell and ('@' in cell or re.search(r'\d', cell)):
                        items.append(cell)
            return items if items else LeadProcessor.parse_input(text)
        except:
            return LeadProcessor.parse_input(text)
    
    @staticmethod
    def process_bulk(items: List[str], session_id: str) -> Dict:
        results = {'valid': [], 'invalid': [], 'total': len(items)}
        conn = sqlite3.connect(Config.DATABASE)
        c = conn.cursor()
        
        for item in items:
            is_email = '@' in item and '.' in item
            is_phone = re.search(r'\d{3}[-.]?\d{3}[-.]?\d{4}', item) or re.search(r'\+?\d{10,}', item)
            
            if is_email:
                valid, normalized, role = Validator.validate_email(item)
                if valid:
                    domain = LeadClassifier.extract_domain(normalized)
                    priority = LeadClassifier.get_priority(role)
                    score = LeadClassifier.get_score(role)
                    
                    results['valid'].append({
                        'value': normalized, 
                        'type': 'email', 
                        'role': role, 
                        'priority': priority,
                        'score': score
                    })
                    
                    c.execute('''INSERT INTO leads (session_id, type, value, domain, role, priority, confidence_score) 
                                 VALUES (?, ?, ?, ?, ?, ?, ?)''',
                             (session_id, 'email', normalized, domain, role, priority, score/100))
                    
                    app.logger.info(f"Saved email: {normalized} | Role: {role} | Priority: {priority}")
                    
            elif is_phone:
                valid, normalized = Validator.validate_phone(item)
                if valid:
                    results['valid'].append({
                        'value': normalized, 
                        'type': 'phone', 
                        'role': 'contact', 
                        'priority': 'medium',
                        'score': 50
                    })
                    c.execute('''INSERT INTO leads (session_id, type, value, domain, role, priority, confidence_score) 
                                 VALUES (?, ?, ?, ?, ?, ?, ?)''',
                             (session_id, 'phone', normalized, 'phone', 'contact', 'medium', 0.85))
            else:
                results['invalid'].append({'value': item, 'reason': 'Invalid email or phone'})
        
        conn.commit()
        
        # Update session status
        c.execute('INSERT OR REPLACE INTO sessions (id, status) VALUES (?, ?)', (session_id, 'validated'))
        conn.commit()
        conn.close()
        
        return results

# ==================== Lead Extractor ====================
class LeadExtractor:
    @staticmethod
    def extract_and_classify(session_id: str) -> Dict:
        """Re-classify all leads in a session (ensures roles are set)"""
        conn = sqlite3.connect(Config.DATABASE)
        c = conn.cursor()
        
        # Get all email leads that need classification
        c.execute('SELECT id, value FROM leads WHERE session_id = ? AND type = "email"', (session_id,))
        leads = c.fetchall()
        
        extracted_count = 0
        for lead_id, value in leads:
            # Detect role
            role = LeadClassifier.detect_role(value)
            priority = LeadClassifier.get_priority(role)
            domain = LeadClassifier.extract_domain(value)
            score = LeadClassifier.get_score(role)
            
            # Update the lead
            c.execute('''UPDATE leads 
                         SET role = ?, priority = ?, domain = ?, confidence_score = ? 
                         WHERE id = ?''',
                     (role, priority, domain, score/100, lead_id))
            extracted_count += 1
            app.logger.info(f"Updated lead {lead_id}: {value} -> Role: {role}, Priority: {priority}")
        
        # Also update phone leads with proper priority
        c.execute('SELECT id FROM leads WHERE session_id = ? AND type = "phone" AND role IS NULL', (session_id,))
        phone_leads = c.fetchall()
        for lead_id in phone_leads:
            c.execute('UPDATE leads SET role = ?, priority = ? WHERE id = ?', ('contact', 'medium', lead_id[0]))
        
        c.execute('UPDATE sessions SET status = ? WHERE id = ?', ('extracted', session_id))
        conn.commit()
        conn.close()
        
        return {'extracted': extracted_count, 'session_id': session_id}

# ==================== Lead Sorter ====================
class LeadSorter:
    @staticmethod
    def sort_by_domain(leads: List[Dict]) -> Dict:
        groups = defaultdict(list)
        for lead in leads:
            domain = lead.get('domain', 'unknown')
            groups[domain].append(lead)
        return dict(sorted(groups.items(), key=lambda x: len(x[1]), reverse=True))
    
    @staticmethod
    def sort_by_role(leads: List[Dict]) -> Dict:
        groups = defaultdict(list)
        for lead in leads:
            role = lead.get('role', 'generic')
            groups[role].append(lead)
        order = ['executive', 'management', 'sales', 'marketing', 'finance', 'legal', 'technical', 'hr', 'support', 'admin', 'contact', 'personal', 'generic']
        return {role: groups[role] for role in order if role in groups}
    
    @staticmethod
    def sort_by_priority(leads: List[Dict]) -> Dict:
        groups = defaultdict(list)
        for lead in leads:
            priority = lead.get('priority', 'low')
            groups[priority].append(lead)
        return {p: groups[p] for p in ['high', 'medium', 'low'] if p in groups}

# ==================== Export Manager ====================
class ExportManager:
    @staticmethod
    def to_csv(leads: List[Dict]) -> str:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=['type', 'value', 'domain', 'role', 'priority', 'score'])
        writer.writeheader()
        for lead in leads:
            writer.writerow({
                'type': lead.get('type', ''),
                'value': lead.get('value', ''),
                'domain': lead.get('domain', ''),
                'role': lead.get('role', ''),
                'priority': lead.get('priority', ''),
                'score': f"{lead.get('confidence_score', 0)*100:.0f}%"
            })
        return output.getvalue()
    
    @staticmethod
    def to_excel(leads: List[Dict]) -> bytes:
        data = []
        for lead in leads:
            data.append({
                'Type': lead.get('type', ''),
                'Value': lead.get('value', ''),
                'Domain': lead.get('domain', ''),
                'Role': lead.get('role', '').upper(),
                'Priority': lead.get('priority', '').upper(),
                'Score': f"{lead.get('confidence_score', 0)*100:.0f}%"
            })
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Leads', index=False)
            # Color code by priority
            worksheet = writer.sheets['Leads']
            for row_idx, lead in enumerate(data, start=2):
                if lead['Priority'] == 'HIGH':
                    for col in range(1, 7):
                        worksheet.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif lead['Priority'] == 'MEDIUM':
                    for col in range(1, 7):
                        worksheet.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def to_json(leads: List[Dict]) -> str:
        return json.dumps({'total': len(leads), 'leads': leads}, indent=2, default=str)

# ==================== API Routes ====================
@app.route('/api/upload-csv', methods=['POST'])
@limiter.limit("100 per hour")
def upload_csv():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    
    content = file.read()
    items = LeadProcessor.parse_csv(content)
    
    if not items:
        return jsonify({'error': 'No valid data found'}), 400
    
    session_id = str(uuid.uuid4())[:8]
    results = LeadProcessor.process_bulk(items, session_id)
    
    return jsonify({
        'success': True, 'session_id': session_id,
        'total': results['total'], 'valid_count': len(results['valid']),
        'invalid_count': len(results['invalid']), 'valid_items': results['valid'],
        'invalid_items': results['invalid'][:20]
    })

@app.route('/api/validate-text', methods=['POST'])
@limiter.limit("100 per hour")
def validate_text():
    data = request.get_json()
    text = data.get('text', '')
    
    if not text:
        return jsonify({'error': 'No text provided'}), 400
    
    items = LeadProcessor.parse_input(text)
    if not items:
        return jsonify({'error': 'No valid items found'}), 400
    
    session_id = str(uuid.uuid4())[:8]
    results = LeadProcessor.process_bulk(items, session_id)
    
    return jsonify({
        'success': True, 'session_id': session_id,
        'total': results['total'], 'valid_count': len(results['valid']),
        'invalid_count': len(results['invalid']), 'valid_items': results['valid'],
        'invalid_items': results['invalid'][:20]
    })

@app.route('/api/extract/<session_id>', methods=['POST'])
def extract_leads(session_id):
    result = LeadExtractor.extract_and_classify(session_id)
    return jsonify({'success': True, **result})

@app.route('/api/sort/<session_id>', methods=['GET'])
def sort_leads(session_id):
    sort_type = request.args.get('type', 'domain')
    
    conn = sqlite3.connect(Config.DATABASE)
    c = conn.cursor()
    c.execute('SELECT type, value, domain, role, priority, confidence_score FROM leads WHERE session_id = ?', (session_id,))
    leads_data = c.fetchall()
    conn.close()
    
    if not leads_data:
        return jsonify({'error': 'No leads found'}), 404
    
    leads = []
    for l in leads_data:
        leads.append({
            'type': l[0], 
            'value': l[1], 
            'domain': l[2] or 'unknown', 
            'role': l[3] or 'generic', 
            'priority': l[4] or 'low', 
            'confidence_score': l[5] or 0.5
        })
    
    if sort_type == 'domain':
        sorted_data = LeadSorter.sort_by_domain(leads)
    elif sort_type == 'role':
        sorted_data = LeadSorter.sort_by_role(leads)
    elif sort_type == 'priority':
        sorted_data = LeadSorter.sort_by_priority(leads)
    else:
        return jsonify({'error': 'Invalid sort type'}), 400
    
    return jsonify({'sort_type': sort_type, 'groups': {k: len(v) for k, v in sorted_data.items()}, 'leads': sorted_data})

@app.route('/api/export/<session_id>', methods=['GET'])
def export_leads(session_id):
    format_type = request.args.get('format', 'csv')
    sort_by = request.args.get('sort_by', 'domain')
    
    conn = sqlite3.connect(Config.DATABASE)
    c = conn.cursor()
    c.execute('SELECT type, value, domain, role, priority, confidence_score FROM leads WHERE session_id = ?', (session_id,))
    leads_data = c.fetchall()
    conn.close()
    
    if not leads_data:
        return jsonify({'error': 'No leads found'}), 404
    
    leads = []
    for l in leads_data:
        leads.append({
            'type': l[0], 
            'value': l[1], 
            'domain': l[2] or 'unknown', 
            'role': l[3] or 'generic', 
            'priority': l[4] or 'low', 
            'confidence_score': l[5] or 0.5
        })
    
    # Apply sorting
    if sort_by == 'domain':
        sorted_groups = LeadSorter.sort_by_domain(leads)
        leads = [lead for group in sorted_groups.values() for lead in group]
    elif sort_by == 'role':
        sorted_groups = LeadSorter.sort_by_role(leads)
        leads = [lead for group in sorted_groups.values() for lead in group]
    elif sort_by == 'priority':
        sorted_groups = LeadSorter.sort_by_priority(leads)
        leads = [lead for group in sorted_groups.values() for lead in group]
    
    if format_type == 'csv':
        content = ExportManager.to_csv(leads)
        return Response(content, mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename=leads_{session_id}.csv'})
    elif format_type == 'excel':
        content = ExportManager.to_excel(leads)
        return Response(content, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                       headers={'Content-Disposition': f'attachment; filename=leads_{session_id}.xlsx'})
    elif format_type == 'json':
        content = ExportManager.to_json(leads)
        return Response(content, mimetype='application/json', headers={'Content-Disposition': f'attachment; filename=leads_{session_id}.json'})
    
    return jsonify({'error': 'Invalid format'}), 400

@app.route('/api/session-status/<session_id>', methods=['GET'])
def session_status(session_id):
    conn = sqlite3.connect(Config.DATABASE)
    c = conn.cursor()
    c.execute('SELECT status FROM sessions WHERE id = ?', (session_id,))
    result = c.fetchone()
    conn.close()
    
    status = result[0] if result else 'none'
    return jsonify({'session_id': session_id, 'status': status})

@app.route('/api/stats', methods=['GET'])
def get_stats():
    conn = sqlite3.connect(Config.DATABASE)
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM sessions')
    total_sessions = c.fetchone()[0] or 0
    c.execute('SELECT COUNT(*) FROM leads')
    total_leads = c.fetchone()[0] or 0
    c.execute('SELECT COUNT(DISTINCT value) FROM leads WHERE type = "email"')
    unique_emails = c.fetchone()[0] or 0
    c.execute('SELECT COUNT(DISTINCT value) FROM leads WHERE type = "phone"')
    unique_phones = c.fetchone()[0] or 0
    c.execute('SELECT COUNT(DISTINCT domain) FROM leads WHERE domain != "unknown" AND domain != "phone"')
    unique_domains = c.fetchone()[0] or 0
    conn.close()
    return jsonify({'total_sessions': total_sessions, 'total_leads': total_leads,
                    'unique_emails': unique_emails, 'unique_phones': unique_phones,
                    'unique_domains': unique_domains})

@app.route('/api/clear-session/<session_id>', methods=['DELETE'])
def clear_session(session_id):
    try:
        conn = sqlite3.connect(Config.DATABASE)
        c = conn.cursor()
        c.execute('DELETE FROM leads WHERE session_id = ?', (session_id,))
        c.execute('DELETE FROM sessions WHERE id = ?', (session_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== Web Interface ====================
@app.route('/')
def home():
    return render_template_string(HTML_TEMPLATE)

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>LeadForge Pro - Lead Extractor & Sorter</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css">
    <style>
        * { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; }
        .gradient-bg { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }
        .btn { transition: all 0.2s ease; cursor: pointer; }
        .btn:hover:not(:disabled) { transform: translateY(-2px); box-shadow: 0 5px 15px rgba(0,0,0,0.2); }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; }
        .btn-active { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; }
        .priority-high { background: #fee2e2; border-left: 4px solid #dc2626; }
        .priority-medium { background: #fef3c7; border-left: 4px solid #f59e0b; }
        .priority-low { background: #e0e7ff; border-left: 4px solid #6366f1; }
        .role-badge { display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: 10px; font-weight: bold; }
        .role-executive { background: #dc2626; color: white; }
        .role-management { background: #f59e0b; color: white; }
        .role-sales { background: #10b981; color: white; }
        .role-marketing { background: #3b82f6; color: white; }
        .role-finance { background: #8b5cf6; color: white; }
        .role-technical { background: #06b6d4; color: white; }
        .role-contact { background: #6b7280; color: white; }
        .fade-in { animation: fadeIn 0.3s ease-in; }
        @keyframes fadeIn { from { opacity: 0; transform: translateY(10px); } to { opacity: 1; transform: translateY(0); } }
    </style>
</head>
<body class="bg-gray-50">
    <nav class="gradient-bg text-white shadow-xl">
        <div class="container mx-auto px-6 py-4">
            <div class="flex justify-between items-center">
                <div class="flex items-center space-x-3">
                    <i class="fas fa-fire fa-2x"></i>
                    <span class="text-2xl font-bold">LeadForge Pro</span>
                </div>
                <div class="text-sm">
                    <i class="fas fa-check-circle mr-1"></i> Role & Priority Detection Active
                </div>
            </div>
        </div>
    </nav>

    <div class="gradient-bg text-white py-8">
        <div class="container mx-auto px-6 text-center">
            <h1 class="text-3xl font-bold mb-2">Lead Extractor & Domain Sorter</h1>
            <p class="text-lg opacity-90">Automatic role detection (CEO, Manager, Sales, etc.) and priority scoring</p>
        </div>
    </div>

    <div class="container mx-auto px-6 py-8">
        <!-- Step 1: Upload Card -->
        <div id="step1Card" class="bg-white rounded-2xl shadow-xl p-8 mb-8 fade-in">
            <h2 class="text-2xl font-bold mb-4"><i class="fas fa-cloud-upload-alt text-purple-600 mr-2"></i>Step 1: Upload Data</h2>
            
            <div class="grid md:grid-cols-2 gap-6">
                <div class="border-2 border-dashed border-gray-300 rounded-lg p-6 text-center hover:border-purple-500 transition">
                    <i class="fas fa-file-csv text-4xl text-gray-400 mb-3"></i>
                    <h3 class="font-bold mb-2">Upload CSV File</h3>
                    <input type="file" id="csvFile" accept=".csv" class="hidden" onchange="updateFileName(this)">
                    <button id="chooseFileBtn" class="bg-gray-600 text-white px-4 py-2 rounded-lg btn">
                        <i class="fas fa-folder-open mr-2"></i>Choose File
                    </button>
                    <div id="fileName" class="text-sm text-gray-500 mt-2"></div>
                    <button id="uploadBtn" disabled class="mt-3 bg-gray-400 text-white px-6 py-2 rounded-lg btn w-full">
                        <i class="fas fa-upload mr-2"></i>Upload & Validate
                    </button>
                </div>
                
                <div class="border-2 border-gray-300 rounded-lg p-6 hover:border-purple-500 transition">
                    <i class="fas fa-file-alt text-4xl text-gray-400 mb-3"></i>
                    <h3 class="font-bold mb-2">Paste Text</h3>
                    <textarea id="textInput" rows="3" class="w-full px-3 py-2 border border-gray-300 rounded-lg focus:outline-none focus:border-purple-500" 
                        placeholder="ceo@company.com, sales@domain.com, +1234567890&#10;john.doe@gmail.com, manager@business.com, 555-123-4567"></textarea>
                    <button id="validateTextBtn" class="mt-3 bg-green-600 text-white px-6 py-2 rounded-lg btn w-full">
                        <i class="fas fa-check-circle mr-2"></i>Validate & Extract
                    </button>
                </div>
            </div>
            
            <div class="mt-6 p-4 bg-blue-50 rounded-lg">
                <h4 class="font-bold text-blue-800 mb-2"><i class="fas fa-info-circle mr-1"></i>Role Detection Examples:</h4>
                <div class="grid grid-cols-2 md:grid-cols-4 gap-2 text-sm">
                    <div><span class="role-badge role-executive">CEO</span> ceo@company.com</div>
                    <div><span class="role-badge role-executive">CFO</span> cfo@company.com</div>
                    <div><span class="role-badge role-management">Manager</span> manager@company.com</div>
                    <div><span class="role-badge role-sales">Sales</span> sales@company.com</div>
                    <div><span class="role-badge role-marketing">Marketing</span> marketing@company.com</div>
                    <div><span class="role-badge role-technical">Engineer</span> engineer@company.com</div>
                    <div><span class="role-badge role-finance">Finance</span> finance@company.com</div>
                    <div><span class="role-badge role-contact">Personal</span> john.doe@gmail.com</div>
                </div>
            </div>
        </div>

        <!-- Step 2: Results Preview -->
        <div id="step2Card" style="display:none;" class="bg-white rounded-2xl shadow-xl p-8 mb-8 fade-in">
            <h2 class="text-2xl font-bold mb-4"><i class="fas fa-list-check text-purple-600 mr-2"></i>Step 2: Validation Results</h2>
            <div id="validationResults" class="mb-4"></div>
            <div class="flex gap-3">
                <button id="extractBtn" disabled class="bg-gray-400 text-white px-6 py-3 rounded-lg btn">
                    <i class="fas fa-magic mr-2"></i>Extract & Classify Roles
                </button>
                <button id="clearResultsBtn" class="bg-gray-600 text-white px-6 py-3 rounded-lg btn">
                    <i class="fas fa-eraser mr-2"></i>Clear Results
                </button>
            </div>
        </div>

        <!-- Step 3: Sort & Export -->
        <div id="step3Card" style="display:none;" class="bg-white rounded-2xl shadow-xl p-8 mb-8 fade-in">
            <h2 class="text-2xl font-bold mb-4"><i class="fas fa-sort-amount-down text-purple-600 mr-2"></i>Step 3: Sort & Export</h2>
            
            <div class="flex gap-3 mb-6 flex-wrap">
                <button id="sortDomainBtn" disabled class="bg-gray-400 text-white px-4 py-2 rounded-lg btn">
                    <i class="fas fa-globe mr-2"></i>Sort by Domain
                </button>
                <button id="sortRoleBtn" disabled class="bg-gray-400 text-white px-4 py-2 rounded-lg btn">
                    <i class="fas fa-user-tie mr-2"></i>Sort by Role
                </button>
                <button id="sortPriorityBtn" disabled class="bg-gray-400 text-white px-4 py-2 rounded-lg btn">
                    <i class="fas fa-flag mr-2"></i>Sort by Priority
                </button>
            </div>
            
            <div class="flex gap-3 mb-6 flex-wrap">
                <button id="exportCsvBtn" disabled class="bg-gray-400 text-white px-4 py-2 rounded-lg btn">
                    <i class="fas fa-file-csv mr-2"></i>Export CSV
                </button>
                <button id="exportExcelBtn" disabled class="bg-gray-400 text-white px-4 py-2 rounded-lg btn">
                    <i class="fas fa-file-excel mr-2"></i>Export Excel
                </button>
                <button id="exportJsonBtn" disabled class="bg-gray-400 text-white px-4 py-2 rounded-lg btn">
                    <i class="fas fa-code mr-2"></i>Export JSON
                </button>
            </div>
            
            <div id="sortedResults" class="max-h-96 overflow-y-auto"></div>
        </div>

        <!-- Statistics -->
        <div class="bg-white rounded-2xl shadow-xl p-8">
            <h2 class="text-2xl font-bold mb-4"><i class="fas fa-chart-bar text-purple-600 mr-2"></i>Statistics</h2>
            <div class="grid md:grid-cols-5 gap-4">
                <div class="text-center p-4 bg-purple-50 rounded-lg"><div id="totalSessions" class="text-2xl font-bold text-purple-600">0</div><div>Sessions</div></div>
                <div class="text-center p-4 bg-blue-50 rounded-lg"><div id="totalLeads" class="text-2xl font-bold text-blue-600">0</div><div>Total Leads</div></div>
                <div class="text-center p-4 bg-green-50 rounded-lg"><div id="totalEmails" class="text-2xl font-bold text-green-600">0</div><div>Emails</div></div>
                <div class="text-center p-4 bg-orange-50 rounded-lg"><div id="totalPhones" class="text-2xl font-bold text-orange-600">0</div><div>Phones</div></div>
                <div class="text-center p-4 bg-red-50 rounded-lg"><div id="totalDomains" class="text-2xl font-bold text-red-600">0</div><div>Domains</div></div>
            </div>
        </div>
    </div>

    <script>
        let currentSessionId = null;
        let currentSortType = 'domain';
        
        // Button elements
        const uploadBtn = document.getElementById('uploadBtn');
        const validateTextBtn = document.getElementById('validateTextBtn');
        const chooseFileBtn = document.getElementById('chooseFileBtn');
        const extractBtn = document.getElementById('extractBtn');
        const clearResultsBtn = document.getElementById('clearResultsBtn');
        const sortDomainBtn = document.getElementById('sortDomainBtn');
        const sortRoleBtn = document.getElementById('sortRoleBtn');
        const sortPriorityBtn = document.getElementById('sortPriorityBtn');
        const exportCsvBtn = document.getElementById('exportCsvBtn');
        const exportExcelBtn = document.getElementById('exportExcelBtn');
        const exportJsonBtn = document.getElementById('exportJsonBtn');
        
        // Load stats
        loadStats();
        setInterval(loadStats, 30000);
        
        async function loadStats() {
            try {
                const response = await fetch('/api/stats');
                const data = await response.json();
                document.getElementById('totalSessions').textContent = data.total_sessions || 0;
                document.getElementById('totalLeads').textContent = data.total_leads || 0;
                document.getElementById('totalEmails').textContent = data.unique_emails || 0;
                document.getElementById('totalPhones').textContent = data.unique_phones || 0;
                document.getElementById('totalDomains').textContent = data.unique_domains || 0;
            } catch(e) { console.error(e); }
        }
        
        function updateFileName(input) {
            const fileName = input.files[0]?.name || '';
            document.getElementById('fileName').innerHTML = `<i class="fas fa-check-circle text-green-600"></i> ${fileName}`;
            if (fileName) {
                uploadBtn.disabled = false;
                uploadBtn.classList.remove('bg-gray-400');
                uploadBtn.classList.add('bg-purple-600', 'btn-active');
            } else {
                uploadBtn.disabled = true;
                uploadBtn.classList.add('bg-gray-400');
                uploadBtn.classList.remove('bg-purple-600', 'btn-active');
            }
        }
        
        async function uploadCSV() {
            const fileInput = document.getElementById('csvFile');
            if (!fileInput.files.length) {
                showNotification('Please select a CSV file', 'error');
                return;
            }
            
            const formData = new FormData();
            formData.append('file', fileInput.files[0]);
            
            showNotification('Uploading and validating...', 'info');
            uploadBtn.disabled = true;
            uploadBtn.innerHTML = '<i class="fas fa-spinner fa-spin mr-2"></i>Processing...';
            
            try {
                const response = await fetch('/api/upload-csv', { method: 'POST', body: formData });
                const data = await response.json();
                
                if (data.success) {
                    currentSessionId = data.session_id;
                    displayValidationResults(data);
                    showNotification(`✅ Valid: ${data.valid_count} | ❌ Invalid: ${data.invalid_count}`, 'success');
                } else {
                    showNotification('Error: ' + data.error, 'error');
                }
            } catch(e) {
                showNotification('Error: ' + e.message, 'error');
            } finally {
                uploadBtn.disabled = false;
                uploadBtn.innerHTML = '<i class="fas fa-upload mr-2"></i>Upload & Validate';
            }
        }
        
        async function validateText() {
            const text = document.getElementById('textInput').value;
            if (!text.trim()) {
                showNotification('Please paste some text', 'error');
                return;
            }
            
            showNotification('Validating and extracting...', 'info');
            validateTextBtn.disabled = true;
            validateTextBtn.innerHTML = '<i class="fas fa-spinner fa-spin mr-2"></i>Processing...';
            
            try {
                const response = await fetch('/api/validate-text', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ text: text })
                });
                const data = await response.json();
                
                if (data.success) {
                    currentSessionId = data.session_id;
                    displayValidationResults(data);
                    showNotification(`✅ Valid: ${data.valid_count} | ❌ Invalid: ${data.invalid_count}`, 'success');
                } else {
                    showNotification('Error: ' + data.error, 'error');
                }
            } catch(e) {
                showNotification('Error: ' + e.message, 'error');
            } finally {
                validateTextBtn.disabled = false;
                validateTextBtn.innerHTML = '<i class="fas fa-check-circle mr-2"></i>Validate & Extract';
            }
        }
        
        function displayValidationResults(data) {
            let html = `<div class="grid grid-cols-2 gap-4 mb-4">
                <div class="bg-green-50 p-4 rounded-lg text-center">
                    <div class="text-2xl font-bold text-green-600">${data.valid_count}</div>
                    <div>Valid Leads Found</div>
                </div>
                <div class="bg-red-50 p-4 rounded-lg text-center">
                    <div class="text-2xl font-bold text-red-600">${data.invalid_count}</div>
                    <div>Invalid Items</div>
                </div>
            </div>`;
            
            if (data.valid_items && data.valid_items.length > 0) {
                html += `<div class="mb-4"><h3 class="font-bold mb-2">✅ Valid Items (with detected roles):</h3><div class="space-y-1 max-h-48 overflow-y-auto">`;
                data.valid_items.forEach(item => {
                    const icon = item.type === 'email' ? '<i class="fas fa-envelope text-blue-600"></i>' : '<i class="fas fa-phone text-green-600"></i>';
                    const roleBadge = item.role ? `<span class="role-badge role-${item.role} ml-2">${item.role.toUpperCase()}</span>` : '';
                    const priorityBadge = item.priority ? `<span class="text-xs ml-2 ${item.priority === 'high' ? 'text-red-600' : item.priority === 'medium' ? 'text-yellow-600' : 'text-blue-600'}">${item.priority}</span>` : '';
                    
                    html += `<div class="flex justify-between items-center p-2 bg-gray-50 rounded">
                                <div>${icon} <span class="font-mono">${escapeHtml(item.value)}</span> ${roleBadge} ${priorityBadge}</div>
                                <button onclick="copyToClipboard('${escapeHtml(item.value)}')" class="text-purple-600 hover:text-purple-800"><i class="fas fa-copy"></i></button>
                            </div>`;
                });
                html += `</div></div>`;
            }
            
            if (data.invalid_items && data.invalid_items.length > 0) {
                html += `<div><h3 class="font-bold mb-2">❌ Invalid Items:</h3><div class="space-y-1 max-h-32 overflow-y-auto">`;
                data.invalid_items.forEach(item => {
                    html += `<div class="flex justify-between items-center p-2 bg-red-50 rounded text-sm">
                                <span>${escapeHtml(item.value)}</span>
                                <span class="text-red-600 text-xs">${item.reason || 'Invalid'}</span>
                            </div>`;
                });
                html += `</div></div>`;
            }
            
            document.getElementById('validationResults').innerHTML = html;
            document.getElementById('step2Card').style.display = 'block';
            document.getElementById('step1Card').style.display = 'none';
            
            // Enable extract button
            extractBtn.disabled = false;
            extractBtn.classList.remove('bg-gray-400');
            extractBtn.classList.add('bg-purple-600', 'btn-active');
        }
        
        async function extractLeads() {
            if (!currentSessionId) return;
            
            showNotification('Extracting and classifying leads by role...', 'info');
            extractBtn.disabled = true;
            extractBtn.innerHTML = '<i class="fas fa-spinner fa-spin mr-2"></i>Classifying...';
            
            try {
                const response = await fetch(`/api/extract/${currentSessionId}`, { method: 'POST' });
                const data = await response.json();
                
                if (data.success) {
                    showNotification(`✅ Classified ${data.extracted} leads with roles and priorities!`, 'success');
                    // Load sorted results
                    await sortBy('domain');
                    document.getElementById('step3Card').style.display = 'block';
                } else {
                    showNotification('Error: ' + data.error, 'error');
                }
            } catch(e) {
                showNotification('Error: ' + e.message, 'error');
            } finally {
                extractBtn.disabled = false;
                extractBtn.innerHTML = '<i class="fas fa-magic mr-2"></i>Extract & Classify Roles';
            }
        }
        
        async function sortBy(type) {
            if (!currentSessionId) return;
            
            currentSortType = type;
            
            // Update active button styling
            [sortDomainBtn, sortRoleBtn, sortPriorityBtn].forEach(btn => {
                btn.classList.remove('btn-active');
                btn.classList.add('bg-gray-600');
            });
            
            let activeBtn;
            if (type === 'domain') activeBtn = sortDomainBtn;
            else if (type === 'role') activeBtn = sortRoleBtn;
            else activeBtn = sortPriorityBtn;
            
            activeBtn.classList.remove('bg-gray-600');
            activeBtn.classList.add('btn-active');
            
            showNotification(`Sorting by ${type}...`, 'info');
            
            try {
                const response = await fetch(`/api/sort/${currentSessionId}?type=${type}`);
                const data = await response.json();
                
                if (response.ok) {
                    displaySortedResults(data);
                    // Enable export buttons
                    enableExportButtons();
                }
            } catch(e) {
                showNotification('Error: ' + e.message, 'error');
            }
        }
        
        function enableExportButtons() {
            [exportCsvBtn, exportExcelBtn, exportJsonBtn].forEach(btn => {
                btn.disabled = false;
                btn.classList.remove('bg-gray-400');
                btn.classList.add('bg-green-600', 'btn-active');
            });
        }
        
        function getRoleBadge(role) {
            const classes = {
                'executive': 'role-executive',
                'management': 'role-management',
                'sales': 'role-sales',
                'marketing': 'role-marketing',
                'finance': 'role-finance',
                'technical': 'role-technical',
                'contact': 'role-contact'
            };
            return `<span class="role-badge ${classes[role] || 'bg-gray-500'}">${role.toUpperCase()}</span>`;
        }
        
        function getPriorityClass(priority) {
            if (priority === 'high') return 'priority-high';
            if (priority === 'medium') return 'priority-medium';
            return 'priority-low';
        }
        
        function displaySortedResults(data) {
            let html = `<div class="mb-3 text-sm text-gray-600"><i class="fas fa-info-circle"></i> Sorted by: <strong>${data.sort_type}</strong> | Total groups: ${Object.keys(data.groups).length}</div>`;
            
            for (const [groupName, leads] of Object.entries(data.leads)) {
                const groupSize = leads.length;
                const groupId = groupName.replace(/[^a-zA-Z0-9]/g, '_');
                
                html += `<div class="mb-3 rounded-lg overflow-hidden">
                            <div class="bg-gray-100 p-3 cursor-pointer hover:bg-gray-200 transition" onclick="toggleGroup('${groupId}')">
                                <div class="flex justify-between items-center">
                                    <div>
                                        <i class="fas fa-chevron-down text-sm mr-2"></i>
                                        <strong>${escapeHtml(groupName)}</strong>
                                        <span class="ml-2 text-sm text-gray-500">(${groupSize} leads)</span>
                                    </div>
                                </div>
                            </div>
                            <div id="group_${groupId}" class="space-y-1 p-2 bg-gray-50">`;
                
                leads.forEach(lead => {
                    const icon = lead.type === 'email' ? '<i class="fas fa-envelope text-blue-600"></i>' : '<i class="fas fa-phone text-green-600"></i>';
                    const roleBadge = lead.role ? getRoleBadge(lead.role) : '';
                    const priorityClass = getPriorityClass(lead.priority);
                    const score = lead.confidence_score ? Math.round(lead.confidence_score * 100) : 0;
                    
                    html += `<div class="${priorityClass} flex justify-between items-center p-2 bg-white rounded shadow-sm">
                                <div>
                                    ${icon} <span class="font-mono text-sm">${escapeHtml(lead.value)}</span>
                                    ${roleBadge}
                                    <span class="text-xs ml-1 ${lead.priority === 'high' ? 'text-red-600' : lead.priority === 'medium' ? 'text-yellow-600' : 'text-blue-600'}">
                                        (${score}%)
                                    </span>
                                </div>
                                <button onclick="copyToClipboard('${escapeHtml(lead.value)}')" class="text-purple-600 hover:text-purple-800">
                                    <i class="fas fa-copy"></i>
                                </button>
                            </div>`;
                });
                
                html += `</div></div>`;
            }
            
            document.getElementById('sortedResults').innerHTML = html;
        }
        
        function toggleGroup(groupId) {
            const element = document.getElementById(`group_${groupId}`);
            if (element) {
                if (element.style.display === 'none') {
                    element.style.display = 'block';
                } else {
                    element.style.display = 'none';
                }
            }
        }
        
        async function exportData(format) {
            if (!currentSessionId) {
                showNotification('No data to export', 'error');
                return;
            }
            window.open(`/api/export/${currentSessionId}?format=${format}&sort_by=${currentSortType}`, '_blank');
            showNotification(`Exporting as ${format.toUpperCase()}...`, 'success');
        }
        
        async function clearResults() {
            if (currentSessionId) {
                await fetch(`/api/clear-session/${currentSessionId}`, { method: 'DELETE' });
            }
            document.getElementById('step2Card').style.display = 'none';
            document.getElementById('step3Card').style.display = 'none';
            document.getElementById('step1Card').style.display = 'block';
            document.getElementById('validationResults').innerHTML = '';
            document.getElementById('textInput').value = '';
            document.getElementById('csvFile').value = '';
            document.getElementById('fileName').innerHTML = '';
            document.getElementById('sortedResults').innerHTML = '';
            currentSessionId = null;
            
            // Reset buttons
            extractBtn.disabled = true;
            extractBtn.classList.add('bg-gray-400');
            extractBtn.classList.remove('bg-purple-600', 'btn-active');
            
            [sortDomainBtn, sortRoleBtn, sortPriorityBtn].forEach(btn => {
                btn.disabled = true;
                btn.classList.add('bg-gray-400');
                btn.classList.remove('btn-active', 'bg-gray-600');
            });
            
            [exportCsvBtn, exportExcelBtn, exportJsonBtn].forEach(btn => {
                btn.disabled = true;
                btn.classList.add('bg-gray-400');
                btn.classList.remove('bg-green-600', 'btn-active');
            });
            
            showNotification('Results cleared', 'success');
        }
        
        function copyToClipboard(text) {
            navigator.clipboard.writeText(text);
            showNotification('Copied: ' + text, 'success');
        }
        
        function showNotification(message, type) {
            const colors = { success: 'bg-green-500', error: 'bg-red-500', info: 'bg-blue-500' };
            const icons = { success: 'check-circle', error: 'exclamation-circle', info: 'info-circle' };
            const notification = document.createElement('div');
            notification.className = `fixed bottom-4 right-4 ${colors[type]} text-white px-6 py-3 rounded-lg shadow-lg z-50 fade-in`;
            notification.innerHTML = `<i class="fas fa-${icons[type]} mr-2"></i>${message}`;
            document.body.appendChild(notification);
            setTimeout(() => notification.remove(), 3000);
        }
        
        function escapeHtml(str) {
            if (!str) return '';
            return String(str).replace(/[&<>]/g, function(m) {
                if (m === '&') return '&amp;';
                if (m === '<') return '&lt;';
                if (m === '>') return '&gt;';
                return m;
            });
        }
        
        // Event listeners
        chooseFileBtn.onclick = () => document.getElementById('csvFile').click();
        uploadBtn.onclick = uploadCSV;
        validateTextBtn.onclick = validateText;
        extractBtn.onclick = extractLeads;
        clearResultsBtn.onclick = clearResults;
        sortDomainBtn.onclick = () => sortBy('domain');
        sortRoleBtn.onclick = () => sortBy('role');
        sortPriorityBtn.onclick = () => sortBy('priority');
        exportCsvBtn.onclick = () => exportData('csv');
        exportExcelBtn.onclick = () => exportData('excel');
        exportJsonBtn.onclick = () => exportData('json');
    </script>
</body>
</html>"""

if __name__ == '__main__':
    print("=" * 60)
    print("🔥 LeadForge Pro - Role & Priority Detection Fixed")
    print("=" * 60)
    print(f"📍 Server: http://127.0.0.1:5000")
    print("=" * 60)
    print("🎯 ROLE DETECTION EXAMPLES:")
    print("   ceo@company.com     → Executive (High Priority)")
    print("   manager@company.com → Management (High Priority)")
    print("   sales@company.com   → Sales (Medium Priority)")
    print("   engineer@company.com → Technical (Low Priority)")
    print("   john@gmail.com      → Personal (Low Priority)")
    print("=" * 60)
    
    app.run(debug=False, host='127.0.0.1', port=5000, threaded=True)